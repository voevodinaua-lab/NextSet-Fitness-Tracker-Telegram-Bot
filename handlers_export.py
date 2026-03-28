import logging
import io
import csv
import json
from collections import Counter
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from telegram import Update, ReplyKeyboardMarkup, InputFile
from telegram.ext import ContextTypes

from database import get_user_trainings
from handlers_statistics import parse_training_datetime
from utils_constants import *

logger = logging.getLogger(__name__)

HEADER_FILL = PatternFill(fill_type="solid", fgColor="DDEBF7")
HEADER_FONT = Font(bold=True)
TITLE_FONT = Font(bold=True, size=14)


def _normalize_sets(sets):
    if sets is None:
        return []
    if isinstance(sets, list):
        return sets
    if isinstance(sets, str):
        try:
            data = json.loads(sets)
            return data if isinstance(data, list) else []
        except json.JSONDecodeError:
            return []
    return []


def filter_trainings_by_period(trainings, period_type: str):
    """Отбор завершённых тренировок по периоду (для выгрузки)."""
    if period_type == "all_time":
        return list(trainings)
    now = datetime.now()
    if period_type == "current_month":
        return [
            t
            for t in trainings
            if parse_training_datetime(t["date_start"]).year == now.year
            and parse_training_datetime(t["date_start"]).month == now.month
        ]
    return list(trainings)


def _collect_exercise_counter(trainings):
    cnt = Counter()
    for training in trainings:
        for exercise in training.get("exercises") or []:
            cnt[exercise.get("name") or "—"] += 1
    return cnt


def _auto_width(ws, min_width=10, max_width=45):
    for col_cells in ws.columns:
        max_len = min_width
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, min(len(str(cell.value)), max_width))
        ws.column_dimensions[col_letter].width = max_len + 1


def generate_excel_report(user_id: int, period_type: str):
    """
    Многостраничный Excel: сводка, список тренировок, детали подходов.
    Удобно открыть в Excel / Google Таблицах (Файл → Импорт).
    Возвращает bytes или None.
    """
    trainings = get_user_trainings(user_id, limit=5000)
    if not trainings:
        return None
    trainings = filter_trainings_by_period(trainings, period_type)
    if not trainings:
        return None

    period_title = (
        "За текущий месяц" if period_type == "current_month" else "За всё время"
    )

    total_ex = 0
    strength_n = 0
    cardio_n = 0
    for training in trainings:
        for exercise in training.get("exercises") or []:
            total_ex += 1
            if exercise.get("is_cardio"):
                cardio_n += 1
            else:
                strength_n += 1

    counter = _collect_exercise_counter(trainings)
    top_exercises = counter.most_common(15)

    wb = Workbook()

    # --- Лист «Сводка» ---
    ws0 = wb.active
    ws0.title = "Сводка"
    ws0["A1"] = "Отчёт NextSet"
    ws0["A1"].font = TITLE_FONT
    ws0["A2"] = f"Период: {period_title}"
    ws0["A3"] = f"Сформировано: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws0["A5"] = "Всего завершённых тренировок"
    ws0["B5"] = len(trainings)
    ws0["A6"] = "Всего записей упражнений (входов в тренировку)"
    ws0["B6"] = total_ex
    ws0["A7"] = "Из них силовых"
    ws0["B7"] = strength_n
    ws0["A8"] = "Из них кардио"
    ws0["B8"] = cardio_n

    ws0["A10"] = "Топ упражнений (сколько раз добавлены в тренировки)"
    ws0["A10"].font = HEADER_FONT
    ws0["A11"] = "Упражнение"
    ws0["B11"] = "Раз"
    ws0["A11"].fill = HEADER_FILL
    ws0["B11"].fill = HEADER_FILL
    ws0["A11"].font = HEADER_FONT
    ws0["B11"].font = HEADER_FONT

    row = 12
    for name, count in top_exercises:
        ws0.cell(row=row, column=1, value=name)
        ws0.cell(row=row, column=2, value=count)
        row += 1

    hint = ws0.cell(
        row=row + 1,
        column=1,
        value=(
            "Как открыть в Google Таблицах: загрузите файл на Google Диск →"
            " ПКМ → Открыть с помощью → Таблицы."
        ),
    )
    hint.alignment = Alignment(wrap_text=True)
    _auto_width(ws0)

    # --- Лист «Тренировки» ---
    ws1 = wb.create_sheet("Тренировки")
    headers = [
        "№",
        "Дата начала",
        "Дата окончания",
        "Комментарий",
        "Замеры при старте",
        "Всего упражнений",
        "Силовых",
        "Кардио",
    ]
    for c, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    for i, training in enumerate(trainings, 1):
        ex_list = training.get("exercises") or []
        s_count = sum(1 for e in ex_list if not e.get("is_cardio"))
        c_count = sum(1 for e in ex_list if e.get("is_cardio"))
        ws1.cell(row=i + 1, column=1, value=i)
        ws1.cell(row=i + 1, column=2, value=training.get("date_start"))
        ws1.cell(row=i + 1, column=3, value=training.get("date_end") or "")
        ws1.cell(row=i + 1, column=4, value=training.get("comment") or "")
        ws1.cell(row=i + 1, column=5, value=training.get("measurements") or "")
        ws1.cell(row=i + 1, column=6, value=len(ex_list))
        ws1.cell(row=i + 1, column=7, value=s_count)
        ws1.cell(row=i + 1, column=8, value=c_count)

    _auto_width(ws1)
    ws1.freeze_panes = "A2"

    # --- Лист «Детали подходов» ---
    ws2 = wb.create_sheet("Детали подходов")
    det_headers = [
        "Дата тренировки",
        "Тип",
        "Упражнение",
        "№ подхода",
        "Вес (кг)",
        "Повторения",
        "Время (мин)",
        "Дистанция (м)",
        "Скорость (км/ч)",
        "Комментарий / детали",
    ]
    for c, h in enumerate(det_headers, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    dr = 2
    for training in trainings:
        tdate = training["date_start"]
        for exercise in training.get("exercises") or []:
            if exercise.get("is_cardio"):
                ws2.cell(row=dr, column=1, value=tdate)
                ws2.cell(row=dr, column=2, value="Кардио")
                ws2.cell(row=dr, column=3, value=exercise.get("name"))
                ws2.cell(row=dr, column=4, value="")
                ws2.cell(row=dr, column=5, value="")
                ws2.cell(row=dr, column=6, value="")
                ws2.cell(row=dr, column=7, value=exercise.get("time_minutes"))
                ws2.cell(row=dr, column=8, value=exercise.get("distance_meters"))
                ws2.cell(row=dr, column=9, value=exercise.get("speed_kmh"))
                ws2.cell(row=dr, column=10, value=exercise.get("details", ""))
                dr += 1
            else:
                sets_list = _normalize_sets(exercise.get("sets"))
                if not sets_list:
                    ws2.cell(row=dr, column=1, value=tdate)
                    ws2.cell(row=dr, column=2, value="Силовое")
                    ws2.cell(row=dr, column=3, value=exercise.get("name"))
                    ws2.cell(row=dr, column=4, value="")
                    ws2.cell(row=dr, column=5, value="")
                    ws2.cell(row=dr, column=6, value="")
                    ws2.cell(row=dr, column=7, value="")
                    ws2.cell(row=dr, column=8, value="")
                    ws2.cell(row=dr, column=9, value="")
                    ws2.cell(row=dr, column=10, value="Нет данных по подходам")
                    dr += 1
                    continue
                for si, set_data in enumerate(sets_list, 1):
                    ws2.cell(row=dr, column=1, value=tdate)
                    ws2.cell(row=dr, column=2, value="Силовое")
                    ws2.cell(row=dr, column=3, value=exercise.get("name"))
                    ws2.cell(row=dr, column=4, value=si)
                    w = set_data.get("weight") if isinstance(set_data, dict) else None
                    r = set_data.get("reps") if isinstance(set_data, dict) else None
                    ws2.cell(row=dr, column=5, value=w)
                    ws2.cell(row=dr, column=6, value=r)
                    ws2.cell(row=dr, column=7, value="")
                    ws2.cell(row=dr, column=8, value="")
                    ws2.cell(row=dr, column=9, value="")
                    ws2.cell(row=dr, column=10, value="")
                    dr += 1

    _auto_width(ws2)
    ws2.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def generate_csv_export(user_id, period_type="all_time"):
    """Построчная выгрузка подходов в CSV (UTF-8)."""
    trainings = get_user_trainings(user_id, limit=5000)
    trainings = filter_trainings_by_period(trainings, period_type)
    if not trainings:
        return None

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(
        [
            "Дата тренировки",
            "Тип упражнения",
            "Название упражнения",
            "№ подхода",
            "Вес (кг)",
            "Повторения",
            "Время (мин)",
            "Дистанция (м)",
            "Скорость (км/ч)",
            "Детали",
        ]
    )

    for training in trainings:
        training_date = training["date_start"]
        for exercise in training.get("exercises") or []:
            if exercise.get("is_cardio"):
                writer.writerow(
                    [
                        training_date,
                        "Кардио",
                        exercise.get("name"),
                        "",
                        "",
                        "",
                        exercise.get("time_minutes", ""),
                        exercise.get("distance_meters", ""),
                        exercise.get("speed_kmh", ""),
                        exercise.get("details", ""),
                    ]
                )
            else:
                sets_list = _normalize_sets(exercise.get("sets"))
                if not sets_list:
                    writer.writerow(
                        [
                            training_date,
                            "Силовое",
                            exercise.get("name"),
                            "",
                            "",
                            "",
                            "",
                            "",
                            "",
                            "Нет подходов",
                        ]
                    )
                else:
                    for si, set_data in enumerate(sets_list, 1):
                        w = set_data.get("weight") if isinstance(set_data, dict) else ""
                        r = set_data.get("reps") if isinstance(set_data, dict) else ""
                        writer.writerow(
                            [
                                training_date,
                                "Силовое",
                                exercise.get("name"),
                                si,
                                w,
                                r,
                                "",
                                "",
                                "",
                                "",
                            ]
                        )

    return output.getvalue()


def _main_menu_keyboard():
    return ReplyKeyboardMarkup(
        [
            ["💪 Начать тренировку", "📊 История тренировок"],
            ["📝 Мои упражнения", "📈 Статистика", "📏 Мои замеры"],
            ["📤 Выгрузка данных", "❓ Помощь"],
        ],
        resize_keyboard=True,
    )


async def show_export_menu(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Меню выгрузки данных"""
    user_id = update.message.from_user.id
    trainings = get_user_trainings(user_id, limit=1)

    stats_text = ""
    if trainings:
        stats_text = "\n💾 В базе есть сохранённые тренировки.\n"

    keyboard = [
        ["📗 Excel — вся история", "📗 Excel — текущий месяц"],
        ["📄 CSV — вся история", "📄 CSV — текущий месяц"],
        ["🔙 Главное меню"],
    ]

    await update.message.reply_text(
        f"📤 Выгрузка отчёта{stats_text}\n"
        "📗 Excel (.xlsx) — сводка, список тренировок и все подходы (удобно для Google Таблиц).\n"
        "📄 CSV — те же детали подходов в текстовом файле.",
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
    )
    return EXPORT_MENU


async def _send_excel(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
    user_id: int,
    period_type: str,
    period_label: str,
) -> int:
    blob = generate_excel_report(user_id, period_type)
    if not blob:
        await update.message.reply_text(
            f"❌ Нет данных для выгрузки ({period_label}).",
            reply_markup=_main_menu_keyboard(),
        )
        return MAIN_MENU

    fname = f"nextset_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    try:
        await update.message.reply_document(
            document=InputFile(io.BytesIO(blob), filename=fname),
            caption=(
                f"📊 Отчёт Excel — {period_label}\n\n"
                "Листы: «Сводка», «Тренировки», «Детали подходов». "
                "В Google Таблицах: Файл → Импорт → Загрузка."
            ),
            reply_markup=_main_menu_keyboard(),
        )
    except Exception as e:
        logger.error("Ошибка отправки Excel: %s", e, exc_info=True)
        await update.message.reply_text(
            "❌ Не удалось сформировать или отправить Excel. Попробуйте позже.",
            reply_markup=_main_menu_keyboard(),
        )
    return MAIN_MENU


async def _send_csv(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
    user_id: int,
    period_type: str,
    period_label: str,
) -> int:
    csv_data = generate_csv_export(user_id, period_type)
    if not csv_data:
        await update.message.reply_text(
            f"❌ Нет данных для выгрузки ({period_label}).",
            reply_markup=_main_menu_keyboard(),
        )
        return MAIN_MENU

    filename = f"nextset_details_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    try:
        with open(filename, "w", encoding="utf-8-sig", newline="") as f:
            f.write(csv_data)
        with open(filename, "rb") as f:
            await update.message.reply_document(
                document=f,
                filename=filename,
                caption=f"📄 Детали подходов (CSV) — {period_label}",
                reply_markup=_main_menu_keyboard(),
            )
    except Exception as e:
        logger.error("Ошибка выгрузки CSV: %s", e, exc_info=True)
        await update.message.reply_text(
            "❌ Ошибка при создании CSV.",
            reply_markup=_main_menu_keyboard(),
        )
    finally:
        import os

        if os.path.exists(filename):
            os.remove(filename)
    return MAIN_MENU


async def handle_export_menu(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Выбор типа и периода выгрузки."""
    text = (update.message.text or "").strip()

    if text == "🔙 Главное меню":
        from handlers_common import start

        return await start(update, context)

    user_id = update.message.from_user.id

    if text == "📗 Excel — вся история":
        return await _send_excel(update, context, user_id, "all_time", "вся история")
    if text == "📗 Excel — текущий месяц":
        return await _send_excel(
            update, context, user_id, "current_month", "текущий месяц"
        )
    if text == "📄 CSV — вся история":
        return await _send_csv(update, context, user_id, "all_time", "вся история")
    if text == "📄 CSV — текущий месяц":
        return await _send_csv(
            update, context, user_id, "current_month", "текущий месяц"
        )

    await update.message.reply_text(
        "❌ Пожалуйста, используйте кнопки меню.",
        reply_markup=ReplyKeyboardMarkup(
            [
                ["📗 Excel — вся история", "📗 Excel — текущий месяц"],
                ["📄 CSV — вся история", "📄 CSV — текущий месяц"],
                ["🔙 Главное меню"],
            ],
            resize_keyboard=True,
        ),
    )
    return EXPORT_MENU
